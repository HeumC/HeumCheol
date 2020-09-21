import openpyxl

wb = openpyxl.load_workbook('test1.xlsx')

# print(wb.sheetnames)

# for sheet in wb:
#     print(sheet.title)

# mySheet = wb.create_sheet('mySheet')
# print(wb.sheetnames)

# sheet3 = wb.get_sheet_by_name('s3')
# sheet3 = wb['myShee t']

ws = wb.active
# print(ws)
# print(ws['A1'])
# print(ws['A1'].value)

# c = ws['B1']
# print('%d行, %d列: %s' %(c.row, c.column, c.value))
# print('单元格%s: %s' %(c.coordinate, c.value))
#
# print(ws.cell(row=1, column=2))
# print(ws.cell(row=1, column=2).value)
#
# for i in range(1, 3):
#     print(ws.cell(row=i, column=2).value)

colC = ws['C']
row6 = ws[2]
col_range = ws['B:C']
row_range = ws[1:2]

# for col in col_range:
#     for cell in col:
#         print(cell.value)
#
# for row in row_range:
#     for cell in row:
#         print(cell.value)

# for row in ws.iter_rows(min_row=1, max_row=2, max_col=2):
#     for cell in row:
#         print(cell.value)

print(tuple(ws.rows))